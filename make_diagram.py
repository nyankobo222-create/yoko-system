from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

wb = Workbook()

# ==========================================
# ヘルパー関数
# ==========================================
def fill(color): return PatternFill("solid", fgColor=color)
def font(bold=False, size=11, color="000000", italic=False):
    return Font(bold=bold, size=size, color=color, italic=italic, name="Meiryo UI")
def align(h="center", v="center", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
def border(style="thin", color="CCCCCC"):
    s = Side(style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)
def thick_border(color="888888"):
    s = Side(style="medium", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def box(ws, row, col, rowspan, colspan, text, bg, fg="FFFFFF", fontsize=11, bold=True):
    ws.merge_cells(start_row=row, start_column=col,
                   end_row=row+rowspan-1, end_column=col+colspan-1)
    for r in range(row, row + rowspan):
        for c in range(col, col + colspan):
            ws.cell(row=r, column=c).fill = fill(bg)
            ws.cell(row=r, column=c).border = thick_border()
    cell = ws.cell(row=row, column=col)
    cell.value = text
    cell.font = font(bold=bold, size=fontsize, color=fg)
    cell.alignment = align()
    cell.fill = fill(bg)
    cell.border = thick_border()

def arrow(ws, row, col, text):
    c = ws.cell(row=row, column=col)
    c.value = text
    c.font = font(bold=True, size=14, color="888888")
    c.alignment = align()

def label(ws, row, col, colspan, text, color="666666", size=8):
    if colspan > 1:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=col+colspan-1)
    c = ws.cell(row=row, column=col)
    c.value = text
    c.font = font(size=size, color=color, italic=True)
    c.alignment = align()

def title_row(ws, row, text, rowspan=1):
    ws.merge_cells(start_row=row, start_column=1, end_row=row+rowspan-1, end_column=9)
    c = ws.cell(row=row, column=1)
    c.value = text
    c.font = font(bold=True, size=16, color="FFFFFF")
    c.fill = fill("1E1E5A")
    c.alignment = align()
    ws.row_dimensions[row].height = 36
    if rowspan > 1:
        ws.row_dimensions[row+1].height = 8


# ==========================================
# Sheet 1: 仕組みの全体図
# ==========================================
ws1 = wb.active
ws1.title = "①仕組みの全体図"

col_widths = [2, 16, 3, 16, 3, 16, 3, 16, 2]
for i, w in enumerate(col_widths, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w
for r in range(1, 60):
    ws1.row_dimensions[r].height = 18

# タイトル
title_row(ws1, 1, "🎯 JC余興クイズシステム ― 仕組みの全体図", rowspan=2)

# ── セクション1: デプロイの仕組み ──
ws1.merge_cells("B4:I4")
c = ws1["B4"]
c.value = "▶ デプロイの仕組み（コードがサーバーに届くまで）"
c.font = font(bold=True, size=10, color="FFFFFF")
c.fill = fill("2D5A8E")
c.alignment = align(h="left")
ws1.row_dimensions[4].height = 22

ws1.row_dimensions[5].height = 6
for r in [6,7,8,9]: ws1.row_dimensions[r].height = 22

box(ws1, 6, 2, 4, 2, "💻 あなたのPC\nコードを編集する場所", "1E5799", fontsize=10)
arrow(ws1, 8, 4, "→")
label(ws1, 9, 4, 1, "git push", size=8)
box(ws1, 6, 5, 4, 2, "🐙 GitHub\nコードの保管庫", "24292E", fontsize=10)
arrow(ws1, 8, 7, "→")
label(ws1, 9, 7, 1, "自動デプロイ", size=8)
box(ws1, 6, 8, 4, 2, "🚂 Railway\nサーバー（24時間稼働）", "7B2D8B", fontsize=10)

ws1.row_dimensions[11].height = 20
ws1.merge_cells("H11:I11")
c = ws1["H11"]
c.value = "↓ URLが発行される"
c.font = font(bold=True, size=10, color="7B2D8B")
c.alignment = align()

ws1.row_dimensions[12].height = 6
ws1.row_dimensions[13].height = 24
ws1.row_dimensions[14].height = 20
ws1.row_dimensions[15].height = 20
box(ws1, 13, 7, 3, 2, "🌐 公開URL\nyoko-system.up.railway.app\n（QRコードで自動表示）", "27AE60", fontsize=9)

ws1.row_dimensions[16].height = 10

# ── セクション2: 2つのロールとゲームフロー ──
ws1.merge_cells("B17:I17")
c = ws1["B17"]
c.value = "▶ 2つのロールとゲームの流れ"
c.font = font(bold=True, size=10, color="FFFFFF")
c.fill = fill("2D5A8E")
c.alignment = align(h="left")
ws1.row_dimensions[17].height = 22

ws1.row_dimensions[18].height = 8
for r in [19,20,21,22,23]: ws1.row_dimensions[r].height = 20

box(ws1, 19, 2, 5, 2, "🏆 代表者\n（スマホ）\nABCDで正解を回答\n速さで得点変動", "C0392B", fontsize=9, bold=False)
box(ws1, 19, 5, 5, 2, "🖥️ スクリーン\n（プロジェクター）\nQR・回答グラフ\nランキング表示", "2980B9", fontsize=9, bold=False)
box(ws1, 19, 8, 5, 2, "👥 会場参加者\n（スマホ）\n正解する代表者を予測\n複数選択可", "6C63FF", fontsize=9, bold=False)

ws1.row_dimensions[24].height = 8

# フェーズフロー
phases = [
    ("B25", "ロビー", "0A0A2A"),
    ("C25", "→", "FFFFFF"),
    ("D25", "プレビュー\n（予測中）", "4A2080"),
    ("E25", "→", "FFFFFF"),
    ("F25", "3・2・1\nカウント", "1A4080"),
    ("G25", "→", "FFFFFF"),
    ("H25", "出題中", "1A5A2A"),
    ("I25", "→", "FFFFFF"),
]
ws1.row_dimensions[25].height = 28

for cell_id, text, bg in phases:
    c = ws1[cell_id]
    c.value = text
    if text == "→":
        c.font = font(bold=True, size=14, color="888888")
        c.alignment = align()
    else:
        c.font = font(bold=True, size=8, color="FFFFFF")
        c.fill = fill(bg)
        c.alignment = align()
        c.border = thick_border()

ws1.row_dimensions[26].height = 18
ws1.merge_cells("B26:I26")
c = ws1["B26"]
c.value = "→  正解発表（スピードランキング）  →  次の問題へ  →  全問終了後：最終ランキング演出"
c.font = font(size=9, color="888888", italic=True)
c.alignment = align(h="left")

ws1.row_dimensions[27].height = 10

# ── セクション3: リアルタイム通信 ──
ws1.merge_cells("B28:I29")
c = ws1["B28"]
c.value = "⚡ Socket.io（リアルタイム通信）：参加者が回答・予測するたびに、即座にスクリーンのグラフが更新されます"
c.font = font(size=9, color="FFFFFF", bold=True)
c.fill = fill("1A1A4A")
c.alignment = align(h="left")
c.border = thick_border("4444AA")
ws1.row_dimensions[29].height = 20


# ==========================================
# Sheet 2: 当日の操作フロー
# ==========================================
ws2 = wb.create_sheet("②当日の操作手順")

col_w2 = [2, 5, 16, 34, 22, 2]
for i, w in enumerate(col_w2, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

title_row(ws2, 1, "🎯 当日の操作手順", rowspan=2)

ws2.row_dimensions[3].height = 22
headers = ["手順", "担当", "操作内容", "ポイント・補足"]
for col, h in enumerate(headers, 2):
    c = ws2.cell(row=3, column=col)
    c.value = h
    c.font = font(bold=True, size=10, color="FFFFFF")
    c.fill = fill("2D5A8E")
    c.alignment = align()
    c.border = border("thin", "FFFFFF")

steps = [
    # (番号, 担当, 操作内容, ポイント, 背景色, 番号色)
    ("①", "司会者",
     "admin.htmlをスマホ or PCで開く\n→ パスワードを入力してログイン",
     "パスワード：ADMIN_PASSWORD（デフォルト: jcadmin2026）\n※Railway環境変数で変更可",
     "EBF5FB", "2980B9"),

    ("②", "司会者",
     "screen.htmlをプロジェクターで\n全画面表示（F11）",
     "ロビー画面が表示され、QRコードと参加者数がリアルタイムに更新される",
     "EBF5FB", "2980B9"),

    ("③", "全員",
     "スクリーンのQRをスマホでスキャン\n→ 名前入力 → ロール選択",
     "代表者：「代表者として参加」\n会場参加者：「会場として参加」",
     "FFF3E0", "E67E22"),

    ("④", "司会者",
     "管理画面でルール説明テキストを\n確認・編集 →「スクリーンに表示」",
     "デフォルトのルール文が入っているので、得点例の数字だけ確認する\n終わったら「閉じる」",
     "EBF5FB", "2980B9"),

    ("⑤", "司会者",
     "「🔧 テスト問題を開始」を押す",
     "参加者に操作方法を説明しながら動作確認\n得点には入らないので自由に試せる",
     "F0E8FF", "7B2D8B"),

    ("⑥", "司会者",
     "問題バンクから問題をクリックして選択\n→「📺 表示」ボタンを押す",
     "会場参加者のスマホに問題が表示される（代表者には見えない）\nスクリーンは「予測中...」アニメになる",
     "FFF9E0", "F39C12"),

    ("⑦", "会場参加者",
     "スマホで「正解しそうな代表者」を\nタップして選択（複数OK）",
     "何人でも選べる。もう一度タップで解除可能\n問題がスタートしたら変更できなくなる",
     "EDE8FF", "6C63FF"),

    ("⑧", "司会者",
     "「⏱ スタート」ボタンを押す",
     "3・2・1のカウントダウン後、全員のスマホに問題＋選択肢が一斉表示\n会場参加者の予測がロックされる",
     "FFF9E0", "F39C12"),

    ("⑨", "代表者",
     "スマホでABCDのどれかをタップ",
     "タップした瞬間にスクリーンの回答グラフが動く\n速く回答するほど得点が高くなる",
     "FEE8E8", "C0392B"),

    ("⑩", "司会者",
     "全員が回答したのを確認 or 時間切れ後\n正解ボタン（A / B / C / D）を押す",
     "確認ダイアログが出るので正解を確認してから押す\nスクリーンに正解＋スピードランキング＋会場的中率が表示",
     "FFF9E0", "F39C12"),

    ("⑪", "司会者",
     "⑥〜⑩を全問分繰り返す",
     "問題番号は自動でインクリメントされる\n途中でトラブルがあればキック機能（✕退ボタン）で対処",
     "E8F8E8", "27AE60"),

    ("⑫", "司会者",
     "「🏆 最終結果を表示」ボタンを押す",
     "スクリーンに代表者・会場参加者それぞれのランキングが表示",
     "F0E8FF", "7B2D8B"),

    ("⑬", "司会者",
     "ランキング演出パネルで\n「▶ 次へ」を押しながら1人ずつ発表",
     "下位から順に1人ずつ名前が出てくる演出\n中間発表と最終発表で代表者・会場参加者それぞれ選べる",
     "F0E8FF", "7B2D8B"),
]

for i, (num, who, action, point, bg, nc) in enumerate(steps):
    row = i + 4
    ws2.row_dimensions[row].height = 40

    c = ws2.cell(row=row, column=2)
    c.value = num
    c.font = font(bold=True, size=13, color="FFFFFF")
    c.fill = fill(nc)
    c.alignment = align()
    c.border = border()

    c = ws2.cell(row=row, column=3)
    c.value = who
    c.font = font(bold=True, size=10, color=nc)
    c.fill = fill(bg)
    c.alignment = align()
    c.border = border()

    c = ws2.cell(row=row, column=4)
    c.value = action
    c.font = font(size=10, color="333333")
    c.fill = fill(bg)
    c.alignment = align(h="left")
    c.border = border()

    c = ws2.cell(row=row, column=5)
    c.value = point
    c.font = font(size=9, color="666666", italic=True)
    c.fill = fill(bg)
    c.alignment = align(h="left")
    c.border = border()


# ==========================================
# Sheet 3: 採点方式の詳細
# ==========================================
ws3 = wb.create_sheet("③採点方式の詳細")

col_w3 = [2, 20, 20, 26, 2]
for i, w in enumerate(col_w3, 1):
    ws3.column_dimensions[get_column_letter(i)].width = w

title_row(ws3, 1, "📊 採点方式の詳細", rowspan=2)

# ── 代表者 ──
ws3.row_dimensions[3].height = 10
ws3.merge_cells("B4:D4")
c = ws3["B4"]
c.value = "🏆 代表者の採点"
c.font = font(bold=True, size=12, color="FFFFFF")
c.fill = fill("C0392B")
c.alignment = align()
ws3.row_dimensions[4].height = 28

ws3.row_dimensions[5].height = 20
for col, h in enumerate(["状況", "得点", "計算式"], 2):
    c = ws3.cell(row=5, column=col)
    c.value = h
    c.font = font(bold=True, size=10, color="FFFFFF")
    c.fill = fill("8B2020")
    c.alignment = align()
    c.border = border("thin", "FFFFFF")

rep_rows = [
    ("正解（即答・残り時間 > 95%）", "1,000 pt", "500 + 500×速さ比"),
    ("正解（残り時間 50% で回答）", "750 pt",   "500 + 500×0.5"),
    ("正解（ギリギリ・残り1秒）",   "約500 pt", "500 + 500×ほぼ0"),
    ("不正解 または 未回答",        "0 pt",     "—"),
]
rep_bg = ["FEE8E8", "FEEDED", "FFF0F0", "F5F5F5"]
for i, (s, p, f_) in enumerate(rep_rows):
    row = 6 + i
    ws3.row_dimensions[row].height = 26
    for col, val in enumerate([s, p, f_], 2):
        c = ws3.cell(row=row, column=col)
        c.value = val
        c.font = font(size=10, bold=(col == 3), color="333333")
        c.fill = fill(rep_bg[i])
        c.alignment = align()
        c.border = border()

ws3.row_dimensions[10].height = 14

# ── 会場参加者 ──
ws3.merge_cells("B11:D11")
c = ws3["B11"]
c.value = "👥 会場参加者の採点"
c.font = font(bold=True, size=12, color="FFFFFF")
c.fill = fill("6C63FF")
c.alignment = align()
ws3.row_dimensions[11].height = 28

ws3.row_dimensions[12].height = 20
for col, h in enumerate(["状況", "得点（デフォルト）", "備考"], 2):
    c = ws3.cell(row=12, column=col)
    c.value = h
    c.font = font(bold=True, size=10, color="FFFFFF")
    c.fill = fill("4A3FCC")
    c.alignment = align()
    c.border = border("thin", "FFFFFF")

aud_rows = [
    ("的中した代表者が1人いた",                   "+300 pt",           "管理画面で変更可（0〜2000 pt）"),
    ("的中した代表者が2人いた",                   "+600 pt",           "的中人数 × 的中ポイント"),
    ("外れた代表者が1人いた",                     "−100 pt",           "管理画面で変更可（0〜2000 pt）"),
    ("外れた代表者が2人・的中0人",                "−200 pt → 0 pt",   "合計がマイナスになる場合は0に切り上げ"),
    ("的中2人・外れ1人（デフォルト）",            "+500 pt",           "300×2 + (−100)×1 = 500 pt"),
    ("未投票（誰も選ばなかった）",                 "0 pt",              "投票なし扱い"),
]
aud_bg = ["EDE8FF", "E8E4FF", "F0EDFF", "F5F2FF", "EAE6FF", "F8F8FF"]
for i, (s, p, r_) in enumerate(aud_rows):
    row = 13 + i
    ws3.row_dimensions[row].height = 26
    for col, val in enumerate([s, p, r_], 2):
        c = ws3.cell(row=row, column=col)
        c.value = val
        c.font = font(size=10, bold=(col == 3), color="333333")
        c.fill = fill(aud_bg[i])
        c.alignment = align()
        c.border = border()

ws3.row_dimensions[19].height = 14

# ── 注意事項 ──
ws3.merge_cells("B20:D22")
c = ws3["B20"]
c.value = ("⚠️ 注意事項\n\n"
           "• 全データはサーバーのメモリ上に保持されます。サーバー再起動でリセットされます\n"
           "• 切断後2分以内の再接続なら、スコアは自動復元されます\n"
           "• ゲームリセット（↺ボタン）を押すと全参加者データが消えます（確認ダイアログあり）")
c.font = font(size=9, color="7D3C00")
c.fill = fill("FEF9E7")
c.alignment = align(h="left")
c.border = thick_border("F39C12")
ws3.row_dimensions[20].height = 22
ws3.row_dimensions[21].height = 22
ws3.row_dimensions[22].height = 22


# ==========================================
# Sheet 4: 費用とメンテナンス
# ==========================================
ws4 = wb.create_sheet("④費用とメンテナンス")

col_w4 = [2, 22, 22, 28, 2]
for i, w in enumerate(col_w4, 1):
    ws4.column_dimensions[get_column_letter(i)].width = w

title_row(ws4, 1, "💰 費用とメンテナンス", rowspan=2)

ws4.row_dimensions[3].height = 10
ws4.row_dimensions[4].height = 22
for col, h in enumerate(["サービス", "費用", "内容"], 2):
    c = ws4.cell(row=4, column=col)
    c.value = h
    c.font = font(bold=True, size=10, color="FFFFFF")
    c.fill = fill("1E5799")
    c.alignment = align()
    c.border = border("thin", "FFFFFF")

costs = [
    ("GitHub",
     "無料",
     "コードの保管・バージョン管理\npushするだけで自動でサーバーに反映される"),
    ("Railway（Hobbyプラン）",
     "月額 約750円\n（$5/月）",
     "サーバーの稼働費\n・イベント当日だけ使えば実質数円〜数十円\n・使わない期間はサービスを停止すれば節約可能"),
    ("合計（フル稼働）",
     "月 約750円\n（年間 約9,000円）",
     "イベント時だけ稼働なら実質もっと安い\nドメイン費用は不要（railway.appのサブドメインを使用）"),
]

row_colors = ["E8F4FD", "F0E8FF", "FFF9C0"]
for i, (svc, cost, desc) in enumerate(costs):
    row = 5 + i
    ws4.row_dimensions[row].height = 52
    bg = row_colors[i]
    for col, val in enumerate([svc, cost, desc], 2):
        c = ws4.cell(row=row, column=col)
        c.value = val
        c.font = font(size=10, bold=(col == 3), color="333333")
        c.fill = fill(bg)
        c.alignment = align()
        c.border = border()

ws4.row_dimensions[8].height = 14
ws4.merge_cells("B9:D11")
c = ws4["B9"]
c.value = ("🔧 メンテナンス方法\n\n"
           "コードを変更したい場合：PC でファイルを編集 → git push するだけで自動でサーバーに反映\n"
           "（手動での再起動は不要。Railway がデプロイを自動実行します）")
c.font = font(size=10, color="1A5276")
c.fill = fill("EBF5FB")
c.alignment = align(h="left")
c.border = thick_border("2980B9")
ws4.row_dimensions[9].height = 22
ws4.row_dimensions[10].height = 22
ws4.row_dimensions[11].height = 22

ws4.row_dimensions[12].height = 14
ws4.merge_cells("B13:D15")
c = ws4["B13"]
c.value = ("🔐 パスワードの変更方法\n\n"
           "Railway Dashboard → プロジェクトを選択 → Variables タブ\n"
           "→「ADMIN_PASSWORD」を追加または変更 → 自動再デプロイで反映")
c.font = font(size=10, color="4A235A")
c.fill = fill("F5EEF8")
c.alignment = align(h="left")
c.border = thick_border("7B2D8B")
ws4.row_dimensions[13].height = 22
ws4.row_dimensions[14].height = 22
ws4.row_dimensions[15].height = 22


# ==========================================
# 保存
# ==========================================
out = r"D:\jc\yoko_system\JC余興クイズシステム_仕組み図解.xlsx"
wb.save(out)
print(f"保存完了: {out}")
